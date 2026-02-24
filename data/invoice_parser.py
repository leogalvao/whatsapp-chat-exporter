"""Invoice parser for Snow Removal Deployment Tracking.

Supports Excel (.xlsx) and CSV (.csv) invoice files with three formats:
1. Simple billing: BUILDING NAME, Ward, Service Area, Billing Street, Price column
2. Pre-treatment report: Snow Priority, Building Name, Billing Street, Ward, crew/completion fields
3. Completion report: Building Name, Billing Street, Ward, crew assignments, timestamps, verification

Auto-detects format from column headers.
"""

import csv
import io
import os
import re
from datetime import datetime

import openpyxl


def _normalize(s):
    if not s:
        return ""
    return re.sub(r"[^a-z0-9 ]", "", str(s).lower().strip())


def _detect_format(headers):
    h_lower = [_normalize(h) for h in headers if h]
    h_str = " ".join(h_lower)

    if "snow  ice removal" in h_str or "snow ice removal" in h_str or "snow melt" in h_str or "pretreatment" in h_str.replace("-", ""):
        if "deployed removal crew" in h_str or "created datetime" in h_str or "percentage completed" in h_str:
            return "completion_report"
        if "deployed pretreatment crew" in h_str or "sidewalks pretreated" in h_str:
            return "pretreatment_report"
        return "simple_billing"

    if "deployed removal crew" in h_str or "contractor snow removal" in h_str:
        return "completion_report"
    if "deployed pretreatment crew" in h_str or "sidewalks pretreated" in h_str:
        return "pretreatment_report"

    for h in h_lower:
        if any(k in h for k in ["snow", "ice", "melt", "removal", "pre treatment", "pretreatment"]):
            return "simple_billing"

    return "simple_billing"


def _detect_deployment_type(headers, title_row=None, filename=None):
    primary = ""
    if title_row:
        primary = str(title_row).lower()
    if filename:
        primary += " " + str(filename).lower()

    if primary:
        if "pre-treatment" in primary or "pretreatment" in primary or "pre treatment" in primary:
            return "Pre-Treatment"
        if "snow removal" in primary:
            return "Snow Removal"
        if "ice removal" in primary:
            return "Ice Removal"
        if "melt only" in primary or "melt application" in primary:
            return "Snow Melt"

    all_text = " ".join([str(h).lower() for h in headers if h])
    if "pre-treatment" in all_text or "pretreatment" in all_text:
        return "Pre-Treatment"
    if "ice removal" in all_text or "ice melt" in all_text:
        return "Ice Removal"
    if "melt only" in all_text or "melt application" in all_text:
        return "Snow Melt"
    if "snow removal" in all_text or "snow & ice" in all_text:
        return "Snow Removal"
    return "Snow Removal"


def _detect_snow_tier(headers, title_row=None):
    all_text = " ".join([str(h).lower() for h in headers if h])
    if title_row:
        all_text += " " + str(title_row).lower()

    if '12"-24"' in all_text or "12 to 24" in all_text or ">12" in all_text:
        return "price_12_to_24in"
    if '6"-12"' in all_text or "6 to 12" in all_text or ">6" in all_text:
        return "price_6_to_12in"
    if '<6"' in all_text or "under 6" in all_text:
        return "price_under_6in"
    if "melt only" in all_text or "melt application" in all_text:
        return "price_melt_only"
    if "pre-treatment" in all_text or "pretreatment" in all_text:
        return "price_melt_only"
    return "price_under_6in"


def _detect_date_from_filename(filename):
    patterns = [
        r"(\d{4})-(\d{1,2})-(\d{1,2})",
        r"(\d{4})(\d{2})(\d{2})",
    ]
    for pat in patterns:
        m = re.search(pat, filename)
        if m:
            try:
                return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def _find_price_col(headers):
    skip_words = {"priority", "id", "crew", "service area", "ward", "building", "street",
                  "billing", "name", "status", "percentage", "created", "verification",
                  "attachment", "cleared", "pretreated", "accessible", "fire", "loading",
                  "dumpster", "exits", "contractor", "only ice", "parking lots", "side walks",
                  "crosswalk", "hydrant", "dock", "trash"}
    exact_skip = {"snow", "ice", "removal"}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_lower = str(h).lower().strip()
        if h_lower in exact_skip:
            continue
        if any(k in h_lower for k in ["snow", "ice", "melt", "removal", "pre-treatment", "pretreatment", "pre treatment"]):
            if not any(s in h_lower for s in skip_words):
                return i
    return None


def _read_rows_from_csv(filepath_or_text, is_text=False):
    if is_text:
        lines = filepath_or_text
    else:
        with open(filepath_or_text, "r", encoding="utf-8-sig") as f:
            lines = f.read()

    reader = csv.reader(io.StringIO(lines))
    all_rows = []
    for row in reader:
        converted = []
        for cell in row:
            cell = cell.strip() if cell else None
            if cell is None or cell == "":
                converted.append(None)
            else:
                try:
                    converted.append(float(cell))
                except ValueError:
                    converted.append(cell)
        all_rows.append(converted)
    return all_rows


def _read_rows_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            rows.append(list(row))
        if rows:
            sheets[sheet_name] = rows
    return sheets


def parse_invoice(filepath, filename=None):
    if filename is None:
        filename = os.path.basename(filepath)

    ext = os.path.splitext(filename or filepath)[1].lower()
    if ext == ".csv":
        all_rows = _read_rows_from_csv(filepath)
        sheets_data = {"Sheet1": all_rows} if all_rows else {}
    else:
        sheets_data = _read_rows_from_excel(filepath)

    results = []
    for sheet_name, all_rows in sheets_data.items():
        if len(all_rows) < 2:
            continue
        results.extend(_parse_sheet_rows(all_rows, sheet_name, filename))
    return results


def _parse_sheet_rows(all_rows, sheet_name, filename):
    results = []

    title_row = None
    header_row_idx = None
    for i, row in enumerate(all_rows):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        non_none_count = len(vals)
        has_building = any("building" in _normalize(str(v)) for v in vals)
        has_billing = any("billing" in _normalize(str(v)) for v in vals)
        has_ward = any("ward" in _normalize(str(v)) for v in vals)

        if (has_building or has_billing) and (has_ward or non_none_count >= 4):
            header_row_idx = i
            break
        elif non_none_count <= 2 and i == 0:
            title_row = str(vals[0]) if vals else None

    if header_row_idx is None:
        return results

    headers = all_rows[header_row_idx]
    fmt = _detect_format(headers)
    dep_type = _detect_deployment_type(headers, title_row, filename)
    snow_tier = _detect_snow_tier(headers, title_row)
    dep_date = _detect_date_from_filename(filename)

    col_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_norm = _normalize(str(h))
        if "building" in h_norm and "name" in h_norm:
            col_map["building_name"] = i
        elif "building name" in h_norm:
            col_map["building_name"] = i
        elif "billing" in h_norm and "street" in h_norm:
            col_map["address"] = i
        elif "ward" in h_norm:
            col_map["ward"] = i
        elif "service area" in h_norm:
            col_map["service_area"] = i
        elif "snow priority" in h_norm:
            col_map["snow_priority"] = i
        elif "percentage completed" in h_norm:
            col_map["pct_completed"] = i
        elif "deployed removal crew" in h_norm:
            col_map["removal_crew"] = i
        elif "deployed pretreatment crew" in h_norm:
            col_map["pretreatment_crew"] = i
        elif "created by" in h_norm:
            col_map["created_by"] = i
        elif "created date" in h_norm or "time created" in h_norm:
            col_map["created_time"] = i
        elif "sidewalks pretreated" in h_norm:
            col_map["sidewalks_pretreated"] = i
        elif "parking lots pretreated" in h_norm:
            col_map["parking_pretreated"] = i
        elif "verification status" in h_norm:
            col_map["verification_status"] = i
        elif "verification comment" in h_norm:
            col_map["verification_comments"] = i
        elif "parking lots cleared" in h_norm or "parking" in h_norm and "cleared" in h_norm:
            col_map["parking_cleared"] = i
        elif "side walks cleared" in h_norm or "sidewalk" in h_norm and "cleared" in h_norm:
            col_map["sidewalks_cleared"] = i
        elif "ice melt" in h_norm and "applied" in h_norm:
            col_map["ice_melt_applied"] = i
        elif "snow  ice removal" in h_norm or "snow ice removal" in h_norm:
            col_map["snow_removal_id"] = i

    price_col = _find_price_col(headers)

    line_items = []
    for row in all_rows[header_row_idx + 1:]:
        if not any(v is not None for v in row):
            continue

        bname_idx = col_map.get("building_name")
        if bname_idx is not None and bname_idx < len(row):
            building_name = str(row[bname_idx]).strip() if row[bname_idx] else ""
        else:
            building_name = str(row[0]).strip() if row and row[0] else ""

        if not building_name:
            continue

        skip_keywords = ["total", "subtotal", "grand total", "t&m", "additional time"]
        if any(k in building_name.lower() for k in skip_keywords):
            continue

        item = {"building_name": building_name}

        addr_idx = col_map.get("address")
        if addr_idx is not None and addr_idx < len(row) and row[addr_idx]:
            item["address"] = str(row[addr_idx]).strip()

        ward_idx = col_map.get("ward")
        if ward_idx is not None and ward_idx < len(row) and row[ward_idx]:
            item["ward"] = str(row[ward_idx]).strip()

        sa_idx = col_map.get("service_area")
        if sa_idx is not None and sa_idx < len(row) and row[sa_idx]:
            item["service_area"] = str(row[sa_idx]).strip()

        if price_col is not None and price_col < len(row) and row[price_col] is not None:
            try:
                item["billed_amount"] = float(row[price_col])
            except (ValueError, TypeError):
                pass

        prio_idx = col_map.get("snow_priority")
        if prio_idx is not None and prio_idx < len(row) and row[prio_idx]:
            try:
                item["snow_priority"] = int(row[prio_idx])
            except (ValueError, TypeError):
                item["snow_priority"] = str(row[prio_idx])

        for field in ["removal_crew", "pretreatment_crew", "created_by", "created_time",
                      "pct_completed", "verification_status", "verification_comments",
                      "sidewalks_pretreated", "parking_pretreated",
                      "sidewalks_cleared", "parking_cleared", "ice_melt_applied"]:
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                val = row[idx]
                if isinstance(val, (int, float)):
                    item[field] = val
                else:
                    item[field] = str(val).strip()

        line_items.append(item)

    if line_items:
        total_billed = sum(it.get("billed_amount", 0) for it in line_items)
        results.append({
            "filename": filename,
            "sheet": sheet_name,
            "format": fmt,
            "deployment_type": dep_type,
            "snow_tier": snow_tier,
            "deployment_date": dep_date,
            "title": title_row,
            "total_billed": round(total_billed, 2),
            "site_count": len(line_items),
            "line_items": line_items,
        })

    return results


def parse_invoice_bytes(content_bytes, filename):
    import tempfile
    ext = os.path.splitext(filename)[1].lower() if filename else ".xlsx"
    if ext == ".csv":
        try:
            text = content_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content_bytes.decode("latin-1")
        all_rows = _read_rows_from_csv(text, is_text=True)
        if not all_rows:
            return []
        sheets_data = {"Sheet1": all_rows}
        results = []
        for sheet_name, rows in sheets_data.items():
            if len(rows) < 2:
                continue
            results.extend(_parse_sheet_rows(rows, sheet_name, filename))
        return results
    else:
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name
        try:
            return parse_invoice(tmp_path, filename)
        finally:
            os.unlink(tmp_path)


def match_invoice_to_deployment(invoice_date_str, deployments_list, tolerance_days=2):
    if not invoice_date_str or not deployments_list:
        return None

    try:
        inv_date = datetime.strptime(invoice_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None

    from datetime import timedelta
    best_match = None
    best_dist = float("inf")

    for dep in deployments_list:
        start = dep["start_date"]
        end = dep["end_date"]
        if start <= inv_date <= end:
            return dep["label"]
        dist = min(abs((inv_date - start).days), abs((inv_date - end).days))
        if dist <= tolerance_days and dist < best_dist:
            best_dist = dist
            best_match = dep["label"]

    return best_match


def reconcile_invoice_with_chat(invoice, chat_locations, pricing_index, normalize_fn,
                                billable_routes=None):
    line_items = invoice.get("line_items", [])
    reconciliation = []

    for item in line_items:
        bname = item.get("building_name", "")
        addr = item.get("address", "")
        billed = item.get("billed_amount", 0)

        bname_norm = normalize_fn(bname)
        addr_norm = normalize_fn(addr)

        in_chat = False
        matched_chat_loc = None
        matched_service_areas = []
        for cloc in chat_locations:
            cloc_norm = normalize_fn(cloc)
            if cloc_norm == bname_norm or cloc_norm == addr_norm:
                in_chat = True
                matched_chat_loc = cloc
                break
            if bname_norm and bname_norm in cloc_norm:
                in_chat = True
                matched_chat_loc = cloc
                break
            if addr_norm and addr_norm in cloc_norm:
                in_chat = True
                matched_chat_loc = cloc
                break

        if in_chat and billable_routes is not None and matched_chat_loc:
            loc_norm = normalize_fn(matched_chat_loc)
            for _, route in billable_routes.iterrows():
                route_loc = normalize_fn(str(route.get("location", "")))
                if route_loc == loc_norm:
                    matched_service_areas.append(route.get("service_area", "Unknown"))

        contract_price = None
        if pricing_index:
            pricing = pricing_index.get(bname_norm) or pricing_index.get(addr_norm)
            if pricing:
                tier = invoice.get("snow_tier", "price_under_6in")
                contract_price = pricing.get(tier, 0)

        status = "ok"
        notes = ""
        if not in_chat:
            status = "not_in_chat"
            notes = "Site invoiced but not found in chat data"
        elif billed and contract_price and abs(billed - contract_price) > 1:
            status = "price_mismatch"
            notes = f"Billed ${billed:.2f} vs contract ${contract_price:.2f}"

        reconciliation.append({
            "building_name": bname,
            "address": addr,
            "billed_amount": billed,
            "contract_price": contract_price,
            "in_chat_data": in_chat,
            "matched_chat_location": matched_chat_loc,
            "service_areas": ", ".join(sorted(set(matched_service_areas))) if matched_service_areas else "",
            "status": status,
            "notes": notes,
        })

    chat_not_invoiced = []
    invoiced_names = set(_normalize(it["building_name"]) for it in line_items)
    invoiced_addrs = set(_normalize(it.get("address", "")) for it in line_items)

    for cloc in chat_locations:
        cloc_norm = normalize_fn(cloc)
        found = cloc_norm in invoiced_names or cloc_norm in invoiced_addrs
        if not found:
            for iname in invoiced_names:
                if iname and iname in cloc_norm:
                    found = True
                    break
            if not found:
                for iaddr in invoiced_addrs:
                    if iaddr and iaddr in cloc_norm:
                        found = True
                        break
        if not found:
            chat_not_invoiced.append(cloc)

    sw_routes = 0
    pl_routes = 0
    if billable_routes is not None and not billable_routes.empty:
        for _, route in billable_routes.iterrows():
            sa = str(route.get("service_area", "")).strip()
            if sa == "Parking Lot":
                pl_routes += 1
            else:
                sw_routes += 1

    return {
        "line_items": reconciliation,
        "sites_invoiced": len(line_items),
        "sites_in_chat": sum(1 for r in reconciliation if r["in_chat_data"]),
        "sites_not_in_chat": sum(1 for r in reconciliation if not r["in_chat_data"]),
        "price_mismatches": sum(1 for r in reconciliation if r["status"] == "price_mismatch"),
        "chat_not_invoiced": chat_not_invoiced,
        "chat_not_invoiced_count": len(chat_not_invoiced),
        "billable_routes_total": sw_routes + pl_routes,
        "sw_routes": sw_routes,
        "pl_routes": pl_routes,
    }
